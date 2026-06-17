import os
import sqlite3
from io import BytesIO
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for, flash, session, abort,
                    send_from_directory, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
PHOTO_FOLDER = os.path.join(BASE_DIR, 'static', 'photos')
DB_PATH = os.path.join(BASE_DIR, 'prashasti.db')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'ppt', 'pptx', 'png', 'jpg', 'jpeg'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
PHOTO_CATEGORIES = ['Coaching', 'Faculty', 'Students & Results', 'Gallery']

app = Flask(__name__)
app.config['SECRET_KEY'] = 'change-this-secret-key-in-production'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PHOTO_FOLDER'] = PHOTO_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25 MB max upload

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PHOTO_FOLDER, exist_ok=True)


# ===================== DATABASE HELPERS =====================

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db_schema():
    db = sqlite3.connect(DB_PATH)
    db.executescript('''
        CREATE TABLE IF NOT EXISTS user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            phone TEXT,
            password_hash TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS study_material (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT,
            material_type TEXT NOT NULL,
            filename TEXT,
            original_filename TEXT,
            link_url TEXT,
            uploaded_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            subject TEXT,
            message TEXT NOT NULL,
            rating INTEGER,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS enquiry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT,
            course TEXT,
            message TEXT,
            source TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS photo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            title TEXT,
            filename TEXT NOT NULL,
            uploaded_at TEXT NOT NULL
        );
    ''')
    db.commit()
    db.close()


def now_str():
    return datetime.utcnow().isoformat(timespec='seconds')


def parse_dt(value):
    """Parse an ISO datetime string back into a datetime object."""
    return datetime.fromisoformat(value)


def build_excel(headers, rows, sheet_title='Data'):
    """Build an in-memory .xlsx file from a list of column headers and row tuples/lists."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    header_fill = PatternFill('solid', start_color='0F1F3D')
    header_align = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append(row)

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 50)

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===================== HELPERS =====================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to continue.', 'error')
            return redirect(url_for('login'))
        db = get_db()
        user = db.execute('SELECT * FROM user WHERE id = ?', (session['user_id'],)).fetchone()
        if not user or not user['is_admin']:
            abort(403)
        return f(*args, **kwargs)
    return decorated


def get_photos_by_category(category, limit=None):
    db = get_db()
    query = 'SELECT * FROM photo WHERE category = ? ORDER BY uploaded_at DESC'
    if limit:
        query += f' LIMIT {int(limit)}'
    return db.execute(query, (category,)).fetchall()


@app.context_processor
def inject_user():
    user = None
    if 'user_id' in session:
        db = get_db()
        row = db.execute('SELECT * FROM user WHERE id = ?', (session['user_id'],)).fetchone()
        if row:
            user = dict(row)
    return dict(current_user=user)


@app.template_filter('fmtdate')
def fmtdate(value, fmt='%d %b %Y, %I:%M %p'):
    try:
        return parse_dt(value).strftime(fmt)
    except (ValueError, TypeError):
        return value


# ===================== PUBLIC ROUTES =====================

@app.route('/')
def home():
    photos = {
        'coaching': get_photos_by_category('Coaching', limit=1),
        'faculty': get_photos_by_category('Faculty', limit=4),
        'students': get_photos_by_category('Students & Results', limit=3),
        'gallery': get_photos_by_category('Gallery', limit=6),
    }
    return render_template('index.html', photos=photos)


@app.route('/enquiry', methods=['POST'])
def submit_enquiry():
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    course = request.form.get('course', '').strip()
    message = request.form.get('message', '').strip()
    source = request.form.get('source', 'website').strip()

    if not name or not phone:
        return {'success': False, 'error': 'Name and phone are required.'}, 400

    db = get_db()
    db.execute(
        'INSERT INTO enquiry (name, phone, email, course, message, source, created_at) '
        'VALUES (?, ?, ?, ?, ?, ?, ?)',
        (name, phone, email, course, message, source, now_str())
    )
    db.commit()

    return {'success': True}


@app.route('/study-material')
def study_material_page():
    db = get_db()
    materials = db.execute('SELECT * FROM study_material ORDER BY uploaded_at DESC').fetchall()

    categories = {}
    for m in materials:
        cat = m['category'] or 'General'
        categories.setdefault(cat, []).append(m)

    return render_template('study_material.html', categories=categories)


@app.route('/material/download/<int:material_id>')
def download_material(material_id):
    db = get_db()
    material = db.execute('SELECT * FROM study_material WHERE id = ?', (material_id,)).fetchone()
    if not material or material['material_type'] != 'file' or not material['filename']:
        abort(404)
    return send_from_directory(app.config['UPLOAD_FOLDER'], material['filename'],
                                 as_attachment=True, download_name=material['original_filename'])


@app.route('/feedback', methods=['POST'])
def submit_feedback():
    name = request.form.get('name', '').strip()
    subject = request.form.get('subject', '').strip()
    message = request.form.get('message', '').strip()
    rating = request.form.get('rating', '0')

    if not name or not message:
        return {'success': False, 'error': 'Name and message are required.'}, 400

    try:
        rating_val = int(rating)
        if rating_val < 1 or rating_val > 5:
            rating_val = None
    except ValueError:
        rating_val = None

    db = get_db()
    db.execute(
        'INSERT INTO feedback (name, subject, message, rating, created_at) VALUES (?, ?, ?, ?, ?)',
        (name, subject, message, rating_val, now_str())
    )
    db.commit()

    return {'success': True}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')

        db = get_db()
        user = db.execute('SELECT * FROM user WHERE email = ?', (email,)).fetchone()

        if user and check_password_hash(user['password_hash'], password) and user['is_admin']:
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['is_admin'] = True
            flash(f"Welcome back, {user['name']}!", 'success')
            return redirect(url_for('admin_dashboard'))

        flash('Invalid email or password.', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('home'))


@app.route('/profile', methods=['GET', 'POST'])
@admin_required
def profile():
    db = get_db()
    user = db.execute('SELECT * FROM user WHERE id = ?', (session['user_id'],)).fetchone()

    if request.method == 'POST':
        name = request.form.get('name', user['name']).strip()
        phone = request.form.get('phone', user['phone'] or '').strip()
        new_password = request.form.get('new_password', '').strip()

        if new_password:
            password_hash = generate_password_hash(new_password)
            db.execute(
                'UPDATE user SET name = ?, phone = ?, password_hash = ? WHERE id = ?',
                (name, phone, password_hash, user['id'])
            )
        else:
            db.execute(
                'UPDATE user SET name = ?, phone = ? WHERE id = ?',
                (name, phone, user['id'])
            )
        db.commit()
        session['user_name'] = name
        flash('Profile updated successfully.', 'success')
        return redirect(url_for('profile'))

    return render_template('profile.html', user=user)


# ===================== ADMIN ROUTES =====================

@app.route('/admin')
@admin_required
def admin_dashboard():
    db = get_db()
    total_materials = db.execute('SELECT COUNT(*) FROM study_material').fetchone()[0]
    total_feedbacks = db.execute('SELECT COUNT(*) FROM feedback').fetchone()[0]
    total_enquiries = db.execute('SELECT COUNT(*) FROM enquiry').fetchone()[0]
    total_photos = db.execute('SELECT COUNT(*) FROM photo').fetchone()[0]

    recent_feedbacks = db.execute(
        'SELECT * FROM feedback ORDER BY created_at DESC LIMIT 5'
    ).fetchall()

    recent_enquiries = db.execute(
        'SELECT * FROM enquiry ORDER BY created_at DESC LIMIT 5'
    ).fetchall()

    return render_template('admin_dashboard.html',
                            total_materials=total_materials,
                            total_feedbacks=total_feedbacks,
                            total_enquiries=total_enquiries,
                            total_photos=total_photos,
                            recent_feedbacks=recent_feedbacks,
                            recent_enquiries=recent_enquiries)


@app.route('/admin/materials', methods=['GET', 'POST'])
@admin_required
def admin_materials():
    db = get_db()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '').strip()
        material_type = request.form.get('material_type')  # 'file' or 'link'

        if not title:
            flash('Title is required.', 'error')
            return redirect(url_for('admin_materials'))

        filename = None
        original_filename = None
        link_url = None

        if material_type == 'file':
            file = request.files.get('file')
            if not file or file.filename == '':
                flash('Please select a file to upload.', 'error')
                return redirect(url_for('admin_materials'))
            if not allowed_file(file.filename):
                flash('File type not allowed.', 'error')
                return redirect(url_for('admin_materials'))

            original_filename = secure_filename(file.filename)
            filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{original_filename}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        elif material_type == 'link':
            link_url = request.form.get('link_url', '').strip()
            if not link_url:
                flash('Please provide a link URL.', 'error')
                return redirect(url_for('admin_materials'))

        else:
            flash('Invalid material type.', 'error')
            return redirect(url_for('admin_materials'))

        db.execute(
            'INSERT INTO study_material '
            '(title, description, category, material_type, filename, original_filename, link_url, uploaded_at) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (title, description, category, material_type, filename, original_filename, link_url, now_str())
        )
        db.commit()
        flash('Study material added successfully.', 'success')
        return redirect(url_for('admin_materials'))

    materials = db.execute('SELECT * FROM study_material ORDER BY uploaded_at DESC').fetchall()
    return render_template('admin_materials.html', materials=materials)


@app.route('/admin/materials/delete/<int:material_id>', methods=['POST'])
@admin_required
def delete_material(material_id):
    db = get_db()
    material = db.execute('SELECT * FROM study_material WHERE id = ?', (material_id,)).fetchone()
    if not material:
        abort(404)

    if material['material_type'] == 'file' and material['filename']:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], material['filename'])
        if os.path.exists(file_path):
            os.remove(file_path)

    db.execute('DELETE FROM study_material WHERE id = ?', (material_id,))
    db.commit()
    flash('Study material deleted.', 'success')
    return redirect(url_for('admin_materials'))


@app.route('/admin/feedbacks')
@admin_required
def admin_feedbacks():
    db = get_db()
    feedbacks = db.execute('SELECT * FROM feedback ORDER BY created_at DESC').fetchall()
    return render_template('admin_feedbacks.html', feedbacks=feedbacks)


@app.route('/admin/feedbacks/export')
@admin_required
def export_feedbacks():
    db = get_db()
    feedbacks = db.execute('SELECT * FROM feedback ORDER BY created_at DESC').fetchall()

    headers = ['Name', 'Subject', 'Rating', 'Message', 'Submitted On']
    rows = []
    for fb in feedbacks:
        rows.append([
            fb['name'], fb['subject'] or '', fb['rating'] or '',
            fb['message'], fmtdate(fb['created_at'])
        ])

    buf = build_excel(headers, rows, sheet_title='Feedback')
    filename = f"feedback_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/feedbacks/delete/<int:feedback_id>', methods=['POST'])
@admin_required
def delete_feedback(feedback_id):
    db = get_db()
    db.execute('DELETE FROM feedback WHERE id = ?', (feedback_id,))
    db.commit()
    flash('Feedback deleted.', 'success')
    return redirect(url_for('admin_feedbacks'))


@app.route('/admin/enquiries')
@admin_required
def admin_enquiries():
    db = get_db()
    enquiries = db.execute('SELECT * FROM enquiry ORDER BY created_at DESC').fetchall()
    return render_template('admin_enquiries.html', enquiries=enquiries)


@app.route('/admin/enquiries/export')
@admin_required
def export_enquiries():
    db = get_db()
    enquiries = db.execute('SELECT * FROM enquiry ORDER BY created_at DESC').fetchall()

    headers = ['Name', 'Phone', 'Email', 'Course', 'Message', 'Source', 'Submitted On']
    rows = []
    for e in enquiries:
        rows.append([
            e['name'], e['phone'], e['email'] or '', e['course'] or '',
            e['message'] or '', e['source'] or '', fmtdate(e['created_at'])
        ])

    buf = build_excel(headers, rows, sheet_title='Enquiries')
    filename = f"enquiries_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/enquiries/delete/<int:enquiry_id>', methods=['POST'])
@admin_required
def delete_enquiry(enquiry_id):
    db = get_db()
    db.execute('DELETE FROM enquiry WHERE id = ?', (enquiry_id,))
    db.commit()
    flash('Enquiry deleted.', 'success')
    return redirect(url_for('admin_enquiries'))


@app.route('/admin/photos', methods=['GET', 'POST'])
@admin_required
def admin_photos():
    db = get_db()

    if request.method == 'POST':
        category = request.form.get('category', '').strip()
        title = request.form.get('title', '').strip()
        file = request.files.get('photo')

        if category not in PHOTO_CATEGORIES:
            flash('Invalid category.', 'error')
            return redirect(url_for('admin_photos'))

        if not file or file.filename == '':
            flash('Please select a photo to upload.', 'error')
            return redirect(url_for('admin_photos'))

        if not allowed_image_file(file.filename):
            flash('Only image files (png, jpg, jpeg, webp) are allowed.', 'error')
            return redirect(url_for('admin_photos'))

        original_filename = secure_filename(file.filename)
        filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{original_filename}"
        file.save(os.path.join(app.config['PHOTO_FOLDER'], filename))

        db.execute(
            'INSERT INTO photo (category, title, filename, uploaded_at) VALUES (?, ?, ?, ?)',
            (category, title, filename, now_str())
        )
        db.commit()
        flash('Photo uploaded successfully.', 'success')
        return redirect(url_for('admin_photos'))

    photos = db.execute('SELECT * FROM photo ORDER BY category, uploaded_at DESC').fetchall()
    grouped = {}
    for p in photos:
        grouped.setdefault(p['category'], []).append(p)

    return render_template('admin_photos.html', grouped=grouped, categories=PHOTO_CATEGORIES)


@app.route('/admin/photos/delete/<int:photo_id>', methods=['POST'])
@admin_required
def delete_photo(photo_id):
    db = get_db()
    photo = db.execute('SELECT * FROM photo WHERE id = ?', (photo_id,)).fetchone()
    if not photo:
        abort(404)

    file_path = os.path.join(app.config['PHOTO_FOLDER'], photo['filename'])
    if os.path.exists(file_path):
        os.remove(file_path)

    db.execute('DELETE FROM photo WHERE id = ?', (photo_id,))
    db.commit()
    flash('Photo deleted.', 'success')
    return redirect(url_for('admin_photos'))


# ===================== ERROR HANDLERS =====================

@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403, message='Access Forbidden'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, message='Page Not Found'), 404


def ensure_db_ready():
    """Create DB schema and default admin if not already present. Safe to call on every startup."""
    if not os.path.exists(DB_PATH):
        init_db_schema()
    else:
        init_db_schema()  # CREATE TABLE IF NOT EXISTS is safe to re-run

    db = sqlite3.connect(DB_PATH)
    admin_email = 'admin@prashastiias.com'
    existing = db.execute('SELECT id FROM user WHERE email = ?', (admin_email,)).fetchone()
    if not existing:
        password_hash = generate_password_hash('admin123')
        db.execute(
            'INSERT INTO user (name, email, phone, password_hash, is_admin, created_at) '
            'VALUES (?, ?, ?, ?, 1, ?)',
            ('Admin', admin_email, '', password_hash, now_str())
        )
        db.commit()
    db.close()


# ===================== CLI: CREATE DB & ADMIN =====================

@app.cli.command('init-db')
def init_db_command():
    """Initialize the database and create a default admin user."""
    ensure_db_ready()
    print('Database initialized.')
    print('Admin login -> email: admin@prashastiias.com | password: admin123')
    print('IMPORTANT: Change this password after first login!')


# Ensure DB is ready whenever the module is imported (works with gunicorn too)
ensure_db_ready()


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
